"""
services/excel_parser.py
========================
Parses SRWB RawData Excel uploads. Produces a structured ParseResult
containing validated rows, type-coerced metrics, statistical anomaly
flags, and conflict markers against the live database.

The only public API you need:
    parser = ExcelParser()
    result = parser.parse(file_obj, db_conn)   # pure read — no DB writes

Design notes
------------
- All DB interaction is read-only (anomaly history + conflict lookup).
- No Flask context is used — this module is framework-agnostic.
- The parser never raises on bad data; every issue is captured as a
  RowIssue with severity "error" or "warning" on the ParsedRow.
  Only truly unrecoverable format problems (wrong file type, missing
  required dimension columns) raise ValueError to the caller.
"""

from __future__ import annotations

import re
import sqlite3
import logging
from dataclasses import dataclass, field
from typing import Any, Optional

import openpyxl

log = logging.getLogger(__name__)

# ── Column map: normalised Excel header  →  DB column name ───────────────
#
# "Normalised" means: lowercased, punctuation stripped, spaces → underscores.
# Add synonyms here whenever the template evolves. The parser never cares
# about capitalisation or minor punctuation differences in the Excel header.
#
COLUMN_MAP: dict[str, str] = {
    # ── Dimensions ──────────────────────────────────────────────────────
    "zone":                       "zone",
    "scheme":                     "scheme",
    "month":                      "month",
    "year":                       "year",

    # ── Customers ───────────────────────────────────────────────────────
    "metered_customers":          "cust_metered",
    "cust_metered":               "cust_metered",
    "disconnected_customers":     "cust_disconnected",
    "cust_disconnected":          "cust_disconnected",
    "active_customers":           "cust_active",
    "cust_active":                "cust_active",
    "active_postpaid":            "cust_postpaid",
    "cust_postpaid":              "cust_postpaid",
    "active_prepaid":             "cust_prepaid",
    "cust_prepaid":               "cust_prepaid",
    "active_post_ind_consumers":  "cust_post_ind",
    "active_prep_ind_consumers":  "cust_prep_ind",
    "active_post_inst_consumers": "cust_post_inst",
    "active_prep_inst_consumers": "cust_prep_inst",
    "active_post_com_consumers":  "cust_post_com",
    "active_prep_com_consumers":  "cust_prep_com",
    "active_prep_cwp":            "cust_prep_cwp",

    # ── New Water Connections ────────────────────────────────────────────
    "nwcs_bf":                    "nwc_bf",
    "nwc_bf":                     "nwc_bf",
    "nwcs_applied":               "nwc_applied",
    "nwcs_applied_postpaid":      "nwc_applied",
    "nwc_applied":                "nwc_applied",
    "nwcs_done_postpaid":         "nwc_done_postpaid",
    "nwc_done_postpaid":          "nwc_done_postpaid",
    "prepaid_meters_installed":   "nwc_prepaid_installed",
    "nwc_prepaid_installed":      "nwc_prepaid_installed",
    "nwcs_done":                  "nwc_done",
    "nwc_done":                   "nwc_done",
    "nwcs_cf":                    "nwc_cf",
    "nwc_cf":                     "nwc_cf",

    # ── Production & NRW ────────────────────────────────────────────────
    "vol_produced":               "vol_produced",
    "volume_produced":            "vol_produced",
    "total_rw":                   "vol_rw",
    "vol_rw":                     "vol_rw",
    "total_nrw":                  "vol_nrw",
    "vol_nrw":                    "vol_nrw",
    "vol_billed_postpaid":        "vol_billed_postpaid",
    "vol_billed_prepaid":         "vol_billed_prepaid",
    "nrw_pct":                    "nrw_pct",
    "nrw_":                       "nrw_pct",   # catches "NRW %" after normalise
    "nrw_percent":                "nrw_pct",
    "nrw_fy_2324":                "nrw_pct",   # legacy header variant

    # ── Billed & Collections ─────────────────────────────────────────────
    "billed_prepaid":             "billed_prepaid",
    "collected_prepaid":          "collected_prepaid",
    "billed_postpaid":            "billed_postpaid",
    "collected_postpaid":         "collected_postpaid",
    "total_billed":               "total_billed",
    "total_collections":          "total_collections",

    # ── Expenses ─────────────────────────────────────────────────────────
    "chemicals":                  "exp_chemicals",
    "exp_chemicals":              "exp_chemicals",
    "electricity":                "exp_electricity",
    "exp_electricity":            "exp_electricity",
    "fuel":                       "exp_fuel",
    "exp_fuel":                   "exp_fuel",
    "maintenance":                "exp_maintenance",
    "exp_maintenance":            "exp_maintenance",
    "staff":                      "exp_staff",
    "exp_staff":                  "exp_staff",
    "wages":                      "exp_wages",
    "exp_wages":                  "exp_wages",
    "other_overhead":             "exp_other",
    "other_overheads":            "exp_other",
    "exp_other":                  "exp_other",
    "total_operating_costs":      "exp_total",
    "exp_total":                  "exp_total",

    # ── Service Charges & Rentals ─────────────────────────────────────────
    "service_charge":             "sc_service_charge",
    "sc_service_charge":          "sc_service_charge",
    "meter_rental":               "sc_meter_rental",
    "sc_meter_rental":            "sc_meter_rental",
    "total_sales":                "total_sales",

    # ── Water Treatment ───────────────────────────────────────────────────
    "chlorine_kg":                "chem_chlorine_kg",
    "chem_chlorine_kg":           "chem_chlorine_kg",
    "aluminium_sulphate_kg":      "chem_aluminium_kg",
    "chem_aluminium_kg":          "chem_aluminium_kg",
    "soda_kg":                    "chem_soda_kg",
    "chem_soda_kg":               "chem_soda_kg",
    "algae_floc_litres":          "chem_algae_floc_l",
    "chem_algae_floc_l":          "chem_algae_floc_l",
    "cost_of_chemicals":          "chem_cost",
    "chem_cost":                  "chem_cost",
    "prdchem_ratio":              "chem_prd_ratio",
    "chem_prd_ratio":             "chem_prd_ratio",
    "costvol_produced_chemicals": "chem_cost_per_vol",
    "chem_cost_per_vol":          "chem_cost_per_vol",

    # ── Connectivity ──────────────────────────────────────────────────────
    "customers_applied_for_new_connection": "conn_applied",
    "customers_applied":          "conn_applied",
    "conn_applied":               "conn_applied",
    "days_taken_to_give_a_quotation": "conn_days_quotation",
    "days_quotation":             "conn_days_quotation",
    "conn_days_quotation":        "conn_days_quotation",
    "customers_fully_paid":       "conn_fully_paid",
    "conn_fully_paid":            "conn_fully_paid",
    "paidup_new_water_applicants": "conn_paid_applicants",
    "conn_paid_applicants":       "conn_paid_applicants",
    "days_taken_to_connect_paid_up_customers": "conn_days_connect",
    "days_to_connect":            "conn_days_connect",
    "conn_days_connect":          "conn_days_connect",
    "time_taken_to_resolve_queries": "conn_queries_days",
    "conn_queries_days":          "conn_queries_days",

    # ── Stuck Meters ─────────────────────────────────────────────────────
    "stuck_meters_bf":            "stuck_bf",
    "stuck_bf":                   "stuck_bf",
    "stuck_meters_new":           "stuck_new",
    "stuck_new":                  "stuck_new",
    "stuck_meters_repaired":      "stuck_repaired",
    "stuck_repaired":             "stuck_repaired",
    "stuck_meters_replaced":      "stuck_replaced",
    "stuck_replaced":             "stuck_replaced",
    "stuck_meters_cf":            "stuck_cf",
    "stuck_cf":                   "stuck_cf",

    # ── Breakdowns ────────────────────────────────────────────────────────
    "total_breakdowns":           "breakdown_total",
    "breakdown_total":            "breakdown_total",

    # ── Pipelines ─────────────────────────────────────────────────────────
    "pipe_32mm":                  "pipe_32mm",
    "32mm":                       "pipe_32mm",
    "pipe_50mm":                  "pipe_50mm",
    "50mm":                       "pipe_50mm",
    "pipe_63mm":                  "pipe_63mm",
    "63mm":                       "pipe_63mm",
    "pipe_90mm":                  "pipe_90mm",
    "90mm":                       "pipe_90mm",
    "pipe_110mm":                 "pipe_110mm",
    "110mm":                      "pipe_110mm",
    "total_dev_lines_done":       "pipe_total",
    "total_dev_lines":            "pipe_total",
    "pipe_total":                 "pipe_total",
}

# Columns that must be present and non-null on every row
REQUIRED_DIMS: tuple[str, ...] = ("zone", "scheme", "month", "year")

# Metric columns where a null value is a hard error (row excluded)
REQUIRED_METRICS: tuple[str, ...] = ("vol_produced",)

# Metrics used in anomaly detection (must exist in the records SELECT below)
ANOMALY_METRICS: tuple[str, ...] = (
    "vol_produced",
    "cust_active",
    "nrw_pct",
    "total_billed",
    "total_collections",
    "nwc_done",
)

# A value is anomalous if it's outside this band relative to 3-month average
ANOMALY_HIGH = 3.0    # 3× average → too high
ANOMALY_LOW  = 0.33   # ⅓ of average → too low

MONTH_NAMES: dict[str, int] = {
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8, "september": 9,
    "october": 10, "november": 11, "december": 12,
    "jan": 1, "feb": 2, "mar": 3, "apr": 4,
    "jun": 6, "jul": 7, "aug": 8, "sep": 9,
    "oct": 10, "nov": 11, "dec": 12,
}


# ── Data classes ──────────────────────────────────────────────────────────

@dataclass
class RowIssue:
    severity: str   # "error" | "warning"
    field: str
    message: str

    def to_dict(self) -> dict:
        return {"severity": self.severity, "field": self.field, "message": self.message}


@dataclass
class ParsedRow:
    row_num:  int
    zone:     str
    scheme:   str
    month:    int
    year:     int
    metrics:  dict[str, Any]
    status:   str = "ok"    # "ok" | "warning" | "error"
    issues:   list[RowIssue] = field(default_factory=list)
    conflict: Optional[dict] = None   # populated by _detect_conflicts

    @property
    def key(self) -> tuple[str, str, int, int]:
        return (self.zone, self.scheme, self.month, self.year)

    def add_issue(self, severity: str, field_name: str, message: str) -> None:
        self.issues.append(RowIssue(severity, field_name, message))
        if severity == "error":
            self.status = "error"
        elif severity == "warning" and self.status == "ok":
            self.status = "warning"

    def to_dict(self) -> dict:
        return {
            "row_num":  self.row_num,
            "zone":     self.zone,
            "scheme":   self.scheme,
            "month":    self.month,
            "year":     self.year,
            "metrics":  self.metrics,
            "status":   self.status,
            "issues":   [i.to_dict() for i in self.issues],
            "conflict": self.conflict,
        }


@dataclass
class ParseResult:
    rows:                    list[ParsedRow] = field(default_factory=list)
    unrecognised_columns:    list[str]       = field(default_factory=list)
    missing_required_columns: list[str]      = field(default_factory=list)
    period_month:            Optional[int]   = None
    period_year:             Optional[int]   = None

    @property
    def importable_rows(self) -> list[ParsedRow]:
        return [r for r in self.rows if r.status != "error"]

    @property
    def error_rows(self) -> list[ParsedRow]:
        return [r for r in self.rows if r.status == "error"]

    @property
    def conflict_rows(self) -> list[ParsedRow]:
        return [r for r in self.importable_rows if r.conflict is not None]

    def to_dict(self) -> dict:
        return {
            "total_rows":              len(self.rows),
            "importable_count":        len(self.importable_rows),
            "error_count":             len(self.error_rows),
            "warning_count":           sum(1 for r in self.rows if r.status == "warning"),
            "conflict_count":          len(self.conflict_rows),
            "period_month":            self.period_month,
            "period_year":             self.period_year,
            "unrecognised_columns":    self.unrecognised_columns,
            "missing_required_columns": self.missing_required_columns,
            "rows":                    [r.to_dict() for r in self.rows],
        }


# ── Core parser class ─────────────────────────────────────────────────────

class ExcelParser:
    """
    One instance per upload. Call parse() once; it returns a ParseResult.
    All DB access is SELECT-only.
    """

    def parse(self, file_obj, db_conn: sqlite3.Connection) -> ParseResult:
        """
        Full pipeline:
          open workbook → find sheet → read & map headers
          → parse rows → coerce types → detect anomalies → detect conflicts
        
        Raises ValueError for fatal format problems (bad file, missing
        required dimension columns). All row-level problems are captured
        as RowIssues — they never raise.
        """
        wb = self._open_workbook(file_obj)
        ws = self._find_data_sheet(wb)
        col_map, unrecognised, missing_required = self._read_headers(ws)

        result = ParseResult(
            unrecognised_columns=unrecognised,
            missing_required_columns=missing_required,
        )

        if missing_required:
            raise ValueError(
                f"Required columns not found in sheet: "
                f"{', '.join(missing_required)}. "
                f"Check the file matches the SRWB template and re-upload."
            )

        for row_num, raw_row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            if all(cell is None for cell in raw_row):
                continue
            parsed = self._parse_row(raw_row, row_num, col_map)
            result.rows.append(parsed)

        self._infer_period(result)
        self._detect_anomalies(result, db_conn)
        self._detect_conflicts(result, db_conn)

        log.info(
            "Parse complete: %d rows total, %d importable, %d errors, %d conflicts",
            len(result.rows),
            len(result.importable_rows),
            len(result.error_rows),
            len(result.conflict_rows),
        )
        return result

    # ── Workbook helpers ──────────────────────────────────────────────────

    def _open_workbook(self, file_obj) -> openpyxl.Workbook:
        try:
            return openpyxl.load_workbook(
                file_obj, read_only=True, data_only=True
            )
        except Exception as exc:
            raise ValueError(f"Cannot open workbook: {exc}") from exc

    def _find_data_sheet(self, wb: openpyxl.Workbook):
        """
        Preferred sheet names in priority order.
        Falls back to the first (active) sheet.
        """
        preferred = ("rawdata", "data", "sheet1", "records")
        lower_map = {name.lower(): name for name in wb.sheetnames}
        for p in preferred:
            if p in lower_map:
                return wb[lower_map[p]]
        return wb.active

    # ── Header processing ─────────────────────────────────────────────────

    @staticmethod
    def _normalize_header(raw: Any) -> str:
        """
        'Vol. Produced (m³)'  →  'vol_produced_m'
        'NRW %'               →  'nrw_'          (trailing _ is harmless)
        'NWCs B/F'            →  'nwcs_bf'

        Steps (order matters):
          1. Encode to ASCII, dropping non-ASCII glyphs (strips ³, ², etc.)
          2. Remove slashes/backslashes without leaving a space ('B/F' → 'BF')
          3. Replace remaining punctuation with a space
          4. Collapse whitespace to underscores
        """
        if raw is None:
            return ""
        s = str(raw).strip().lower()
        s = s.encode("ascii", "ignore").decode("ascii")  # drop ³, ², etc.
        s = re.sub(r"[/\\]", "", s)                      # /\ → nothing
        s = re.sub(r"[^\w\s]", " ", s)                   # punctuation → space
        s = re.sub(r"\s+", "_", s.strip())
        s = re.sub(r"_+", "_", s)
        return s.strip("_")

    def _read_headers(
        self, ws
    ) -> tuple[dict[int, str], list[str], list[str]]:
        """
        Returns:
          col_map           – {col_index: db_column_name}
          unrecognised      – raw header strings not in COLUMN_MAP
          missing_required  – required DB dim columns absent from the sheet
        """
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))

        col_map: dict[int, str] = {}
        unrecognised: list[str] = []

        for idx, cell in enumerate(header_row):
            if cell is None:
                continue
            normalised = self._normalize_header(cell)
            if normalised in COLUMN_MAP:
                db_col = COLUMN_MAP[normalised]
                # Last mapping wins if duplicate headers exist
                col_map[idx] = db_col
            else:
                unrecognised.append(str(cell).strip())

        mapped_db_cols = set(col_map.values())
        missing_required = [d for d in REQUIRED_DIMS if d not in mapped_db_cols]

        return col_map, unrecognised, missing_required

    # ── Row parsing ───────────────────────────────────────────────────────

    def _parse_row(
        self,
        raw_row: tuple,
        row_num: int,
        col_map: dict[int, str],
    ) -> ParsedRow:
        # Extract every mapped column from the raw tuple
        raw: dict[str, Any] = {
            db_col: (raw_row[idx] if idx < len(raw_row) else None)
            for idx, db_col in col_map.items()
        }

        # Parse dimensions first
        zone   = self._coerce_str(raw.get("zone"))
        scheme = self._coerce_str(raw.get("scheme"))
        month  = self._coerce_month(raw.get("month"))
        year   = self._coerce_year(raw.get("year"))

        # Separate metrics from dimensions
        metrics: dict[str, Any] = {
            k: v for k, v in raw.items() if k not in REQUIRED_DIMS
        }

        parsed = ParsedRow(
            row_num=row_num,
            zone=zone or "",
            scheme=scheme or "",
            month=month or 0,
            year=year or 0,
            metrics=metrics,
        )

        # ── Validate dimensions ──────────────────────────────────────────
        if not zone:
            parsed.add_issue("error", "zone", "Zone is missing or blank.")
        if not scheme:
            parsed.add_issue("error", "scheme", "Scheme is missing or blank.")
        if month is None:
            parsed.add_issue(
                "error", "month",
                f"Cannot parse month from value: '{raw.get('month')}'."
            )
        if year is None:
            parsed.add_issue(
                "error", "year",
                f"Cannot parse year from value: '{raw.get('year')}'."
            )

        # ── Validate required metrics ────────────────────────────────────
        for req in REQUIRED_METRICS:
            val = metrics.get(req)
            if val is None:
                parsed.add_issue(
                    "error", req,
                    f"'{req}' is required but missing for "
                    f"{zone or '?'} / {scheme or '?'}."
                )

        # ── Coerce all metric values to float (or None) ──────────────────
        for col, val in list(metrics.items()):
            if val is None:
                continue
            coerced = self._coerce_numeric(val)
            if coerced is None and str(val).strip() not in ("", "-", "—"):
                parsed.add_issue(
                    "warning", col,
                    f"Non-numeric value '{val}' for column '{col}' — "
                    f"treated as missing."
                )
            metrics[col] = coerced

        return parsed

    # ── Type coercion helpers ─────────────────────────────────────────────

    @staticmethod
    def _coerce_str(val: Any) -> Optional[str]:
        if val is None:
            return None
        s = str(val).strip()
        return s if s else None

    @staticmethod
    def _coerce_numeric(val: Any) -> Optional[float]:
        if val is None:
            return None
        if isinstance(val, (int, float)):
            return float(val)
        cleaned = re.sub(r"[,\s]", "", str(val).strip())
        try:
            return float(cleaned)
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _coerce_month(val: Any) -> Optional[int]:
        """Accept integer 1–12, month names, or Excel date objects."""
        if val is None:
            return None
        if hasattr(val, "month"):           # datetime from openpyxl
            return val.month
        if isinstance(val, (int, float)):
            m = int(val)
            return m if 1 <= m <= 12 else None
        s = str(val).strip().lower()
        if s.isdigit():
            m = int(s)
            return m if 1 <= m <= 12 else None
        return MONTH_NAMES.get(s)

    @staticmethod
    def _coerce_year(val: Any) -> Optional[int]:
        """Accept 2024 or 24 (→ 2024), or Excel date objects."""
        if val is None:
            return None
        if hasattr(val, "year"):
            return val.year
        try:
            y = int(float(str(val).strip()))
            return 2000 + y if y < 100 else y
        except (ValueError, TypeError):
            return None

    # ── Period inference ──────────────────────────────────────────────────

    @staticmethod
    def _infer_period(result: ParseResult) -> None:
        """Detect the dominant (month, year) across all parsed rows."""
        months = [r.month for r in result.rows if r.month and r.month > 0]
        years  = [r.year  for r in result.rows if r.year  and r.year  > 0]
        if months:
            result.period_month = max(set(months), key=months.count)
        if years:
            result.period_year = max(set(years), key=years.count)

    # ── Anomaly detection ─────────────────────────────────────────────────

    def _detect_anomalies(
        self, result: ParseResult, db_conn: sqlite3.Connection
    ) -> None:
        """
        For each ANOMALY_METRIC, compare the incoming value against the
        trailing 3-month average for that (zone, scheme) in the DB.
        Flag if outside the [ANOMALY_LOW, ANOMALY_HIGH] band.

        Rows with fewer than 2 historical records are skipped — not enough
        history to establish a meaningful baseline.
        """
        cur = db_conn.cursor()
        metric_cols = ", ".join(ANOMALY_METRICS)

        for row in result.importable_rows:
            cur.execute(
                f"""
                SELECT {metric_cols}
                FROM   records
                WHERE  zone   = ?
                  AND  scheme = ?
                  AND  (year < ? OR (year = ? AND month < ?))
                ORDER  BY year DESC, month DESC
                LIMIT  3
                """,
                (row.zone, row.scheme, row.year, row.year, row.month),
            )
            history = cur.fetchall()
            if len(history) < 2:
                continue

            for m_idx, metric in enumerate(ANOMALY_METRICS):
                new_val = row.metrics.get(metric)
                if not isinstance(new_val, (int, float)):
                    continue

                hist_vals = [
                    h[m_idx] for h in history
                    if h[m_idx] is not None
                ]
                if not hist_vals:
                    continue

                avg = sum(hist_vals) / len(hist_vals)
                if avg == 0:
                    continue

                ratio = new_val / avg

                if ratio > ANOMALY_HIGH:
                    row.add_issue(
                        "warning", metric,
                        f"Value {new_val:,.1f} is {ratio:.1f}× the "
                        f"{len(hist_vals)}-month trailing average "
                        f"({avg:,.1f}). Verify before importing."
                    )
                elif ratio < ANOMALY_LOW:
                    row.add_issue(
                        "warning", metric,
                        f"Value {new_val:,.1f} is only {ratio:.0%} of the "
                        f"{len(hist_vals)}-month trailing average "
                        f"({avg:,.1f}). Possible data-entry error."
                    )

    # ── Conflict detection ────────────────────────────────────────────────

    def _detect_conflicts(
        self, result: ParseResult, db_conn: sqlite3.Connection
    ) -> None:
        """
        Batch-fetch all (zone, scheme, month, year) combos that already
        exist in the DB for the importable rows. Attach a conflict dict to
        each row that has an existing record so the caller can surface the
        old vs new values for the operator to resolve.
        """
        importable = result.importable_rows
        if not importable:
            return

        cur = db_conn.cursor()

        # Build a single batch query rather than N individual queries
        placeholders = ", ".join("(?, ?, ?, ?)" for _ in importable)
        params: list[Any] = []
        for row in importable:
            params.extend([row.zone, row.scheme, row.month, row.year])

        cur.execute(
            f"""
            SELECT zone, scheme, month, year,
                   cust_active, vol_produced, nrw_pct, total_billed,
                   total_collections
            FROM   records
            WHERE  (zone, scheme, month, year) IN ({placeholders})
            """,
            params,
        )

        # Key the results for O(1) lookup
        existing: dict[tuple, dict] = {
            (r[0], r[1], r[2], r[3]): {
                "cust_active":       r[4],
                "vol_produced":      r[5],
                "nrw_pct":           r[6],
                "total_billed":      r[7],
                "total_collections": r[8],
            }
            for r in cur.fetchall()
        }

        for row in importable:
            ex = existing.get(row.key)
            if ex:
                row.conflict = {
                    "existing": ex,
                    "incoming": {
                        k: row.metrics.get(k)
                        for k in (
                            "cust_active", "vol_produced", "nrw_pct",
                            "total_billed", "total_collections",
                        )
                    },
                }
